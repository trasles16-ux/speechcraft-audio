"""
Script Handler Module for SpeechCraft

Reads and parses audio description scripts from:
- Excel files (.xlsx)
- SRT subtitle files (.srt)

Supports automatic line placement for audio descriptions.
"""

import re
from dataclasses import dataclass
from typing import List, Optional, Tuple
from pathlib import Path


@dataclass
class ScriptLine:
    """Represents a single line in an audio description script"""
    line_number: int
    time_in_ms: int          # Start time in milliseconds
    time_out_ms: int         # End time in milliseconds
    description: str         # The text to be recorded
    original_line: str       # Original line for reference
    
    @property
    def duration_ms(self) -> int:
        """Get duration in milliseconds"""
        return self.time_out_ms - self.time_in_ms


class ScriptParser:
    """Parse audio description scripts from various formats"""
    
    @staticmethod
    def parse_srt(file_path: str) -> List[ScriptLine]:
        """Parse SRT subtitle file format.
        
        SRT format:
        1
        00:00:05,000 --> 00:00:10,000
        Description text here
        
        Args:
            file_path: Path to .srt file
            
        Returns:
            List of ScriptLine objects
        """
        lines = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as f:
                content = f.read()
        
        # Split into subtitle blocks
        blocks = content.strip().split('\n\n')
        for block in blocks:
            block_lines = block.strip().split('\n')

            if len(block_lines) < 3:
                continue

            try:
                # Line 1: sequence number (must be a plain integer)
                seq_num_str = block_lines[0].strip()
                # Reject malformed SRT blocks where line 1 is not a number
                # (e.g. continuation lines from wrapped descriptions)
                if not seq_num_str.isdigit():
                    continue
                seq_num = int(seq_num_str)

                # Line 2: timecode
                timecode_line = block_lines[1].strip()
                if ' --> ' not in timecode_line:
                    print(f"Warning: Missing timecode in SRT block {seq_num}")
                    continue
                
                time_in_ms = ScriptParser._srt_time_to_ms(time_in_str.strip())
                time_out_ms = ScriptParser._srt_time_to_ms(time_out_str.strip())
                
                # Lines 3+: description
                description = ' '.join(block_lines[2:]).strip()
                
                if description:  # Only add if there's actual text
                    lines.append(ScriptLine(
                        line_number=seq_num,
                        time_in_ms=time_in_ms,
                        time_out_ms=time_out_ms,
                        description=description,
                        original_line=description
                    ))
            except (ValueError, IndexError) as e:
                print(f"Warning: Could not parse SRT block: {e}")
                continue
        
        return lines
    
    @staticmethod
    def parse_excel(file_path: str) -> List[ScriptLine]:
        """Parse Excel spreadsheet with audio description script.
        
        Expected columns:
        - Column A (or "Time In"): Start time in HH:MM:SS.mmm or MM:SS.mmm
        - Column B (or "Time Out"): End time
        - Column C (or "Description"): Text to be recorded
        
        Args:
            file_path: Path to .xlsx file
            
        Returns:
            List of ScriptLine objects
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl not installed. Install with: pip install openpyxl"
            )
        
        lines = []
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Find header row and column indices
        time_in_col = None
        time_out_col = None
        desc_col = None
        
        # Check first row for headers
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header = str(cell.value).lower().strip()
                if 'time' in header and 'in' in header:
                    time_in_col = col_idx
                elif 'time' in header and 'out' in header:
                    time_out_col = col_idx
                elif 'description' in header or 'text' in header or 'ad' in header:
                    desc_col = col_idx
        
        # If headers not found, assume columns A, B, C
        if time_in_col is None:
            time_in_col = 1
        if time_out_col is None:
            time_out_col = 2
        if desc_col is None:
            desc_col = 3
        
        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            try:
                # Get cell values
                time_in_cell = row[time_in_col - 1]
                time_out_cell = row[time_out_col - 1]
                desc_cell = row[desc_col - 1]
                
                if not desc_cell.value:
                    continue
                
                # Parse time values
                time_in_str = str(time_in_cell.value).strip()
                time_out_str = str(time_out_cell.value).strip()
                description = str(desc_cell.value).strip()
                
                time_in_ms = ScriptParser._time_to_ms(time_in_str)
                time_out_ms = ScriptParser._time_to_ms(time_out_str)
                
                if time_in_ms is not None and time_out_ms is not None:
                    lines.append(ScriptLine(
                        line_number=row_idx - 1,
                        time_in_ms=time_in_ms,
                        time_out_ms=time_out_ms,
                        description=description,
                        original_line=description
                    ))
            except (ValueError, AttributeError) as e:
                print(f"Warning: Could not parse row {row_idx}: {e}")
                continue
        
        wb.close()
        return lines
    
    @staticmethod
    def _srt_time_to_ms(time_str: str) -> int:
        """Convert SRT time format (HH:MM:SS,mmm) to milliseconds.
        
        Args:
            time_str: Time string like "00:00:05,000"
            
        Returns:
            Time in milliseconds
        """
        time_str = time_str.replace(',', '.')
        parts = time_str.split(':')
        
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = float(parts[2])
        
        total_ms = (hours * 3600 + minutes * 60 + seconds) * 1000
        return int(total_ms)
    
    @staticmethod
    def _time_to_ms(time_str: str) -> Optional[int]:
        """Convert time string to milliseconds.
        
        Supports formats:
        - HH:MM:SS.mmm
        - MM:SS.mmm
        - MM:SS
        - HH:MM:SS
        
        Args:
            time_str: Time string
            
        Returns:
            Time in milliseconds, or None if invalid
        """
        if not time_str or time_str.lower() == 'none':
            return None
        
        # Clean up
        time_str = time_str.strip().replace(',', '.')
        
        try:
            # Try different patterns
            parts = time_str.split(':')
            
            if len(parts) == 3:
                # HH:MM:SS.mmm
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = float(parts[2])
            elif len(parts) == 2:
                # MM:SS.mmm
                hours = 0
                minutes = int(parts[0])
                seconds = float(parts[1])
            else:
                return None
            
            total_ms = (hours * 3600 + minutes * 60 + seconds) * 1000
            return int(total_ms)
        except (ValueError, IndexError):
            return None


class ScriptValidator:
    """Validate and check script files"""
    
    @staticmethod
    def validate_script(script_lines: List[ScriptLine]) -> Tuple[bool, str]:
        """Validate script for issues.
        
        Args:
            script_lines: List of ScriptLine objects
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not script_lines:
            return False, "Script file is empty"
        
        # Check for overlapping times
        for i, line1 in enumerate(script_lines):
            for line2 in script_lines[i+1:]:
                if line1.time_in_ms < line2.time_out_ms and line2.time_in_ms < line1.time_out_ms:
                    return False, f"Lines {line1.line_number} and {line2.line_number} have overlapping times"
        
        # Check for reasonable durations
        for line in script_lines:
            duration = line.duration_ms
            if duration <= 0:
                return False, f"Line {line.line_number} has invalid duration"
            if duration > 600000:  # More than 10 minutes
                return False, f"Line {line.line_number} has suspiciously long duration"
        
        return True, ""
    
    @staticmethod
    def suggest_fixes(script_lines: List[ScriptLine]) -> List[str]:
        """Suggest fixes for common script issues.
        
        Args:
            script_lines: List of ScriptLine objects
            
        Returns:
            List of suggestions
        """
        suggestions = []
        
        # Check for very short lines
        short_lines = [line for line in script_lines if line.duration_ms < 1000]
        if short_lines:
            suggestions.append(f"{len(short_lines)} lines are shorter than 1 second")
        
        # Check for very long lines
        long_lines = [line for line in script_lines if line.duration_ms > 120000]
        if long_lines:
            suggestions.append(f"{len(long_lines)} lines are longer than 2 minutes")
        
        # Check for very short descriptions (might be typos)
        short_desc = [line for line in script_lines if len(line.description) < 5]
        if short_desc:
            suggestions.append(f"{len(short_desc)} lines have very short descriptions (< 5 chars)")
        
        return suggestions


class ScriptUtils:
    """Utility functions for script handling"""
    
    @staticmethod
    def get_file_format(file_path: str) -> Optional[str]:
        """Detect file format from extension.
        
        Args:
            file_path: Path to file
            
        Returns:
            'srt', 'excel', or None if unknown format
        """
        ext = Path(file_path).suffix.lower()
        
        if ext == '.srt':
            return 'srt'
        elif ext in ['.xlsx', '.xls']:
            return 'excel'
        
        return None
    
    @staticmethod
    def load_script(file_path: str) -> List[ScriptLine]:
        """Load script from any supported format.
        
        Args:
            file_path: Path to script file
            
        Returns:
            List of ScriptLine objects
            
        Raises:
            ValueError: If format not recognized or file invalid
        """
        file_format = ScriptUtils.get_file_format(file_path)
        
        if file_format == 'srt':
            return ScriptParser.parse_srt(file_path)
        elif file_format == 'excel':
            return ScriptParser.parse_excel(file_path)
        else:
            raise ValueError(
                f"Unknown file format. Supported: .srt, .xlsx"
            )
    
    @staticmethod
    def ms_to_time_str(ms: int) -> str:
        """Convert milliseconds to HH:MM:SS format.
        
        Args:
            ms: Time in milliseconds
            
        Returns:
            Time string like "00:05:30"
        """
        total_seconds = ms // 1000
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    @staticmethod
    def ms_to_srt_time_str(ms: int) -> str:
        """Convert milliseconds to SRT time format.
        
        Args:
            ms: Time in milliseconds
            
        Returns:
            Time string like "00:05:30,000"
        """
        total_seconds = ms / 1000
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = total_seconds % 60
        
        return f"{hours:02d}:{minutes:02d}:{seconds:06.3f}".replace('.', ',')
